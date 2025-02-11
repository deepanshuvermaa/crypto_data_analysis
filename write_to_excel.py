from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

def write_to_excel(df, filename="crypto_data.xlsx"):
    try:
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name="Crypto Data")

        
        wb = load_workbook(filename)
        ws = wb["Crypto Data"]

        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        
        column_widths = {
            "A": 20,  # Name
            "B": 10,  # Symbol
            "C": 15,  # Current Price
            "D": 20,  # Market Cap
            "E": 20,  # Total Volume
            "F": 25   # 24-hour Price Change
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                if cell.value > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for positive
                elif cell.value < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for negative

        #bar chart or market cap
        chart = BarChart()
        chart.title = "Market Cap Distribution"
        chart.x_axis.title = "Cryptocurrency"
        chart.y_axis.title = "Market Cap (USD)"
        data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row, max_col=4)
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "H2")

        # Save the workbook
        wb.save(filename)
        print("Data written and formatted in Excel successfully!")

    except Exception as e:
        print(f"Failed to write to Excel: {e}")
