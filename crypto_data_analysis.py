import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import time
import os

# Display full DataFrame output without truncation
pd.set_option('display.max_columns', None)
pd.set_option('display.expand_frame_repr', False)

# API URL and parameters to fetch cryptocurrency data
api_url = "https://api.coingecko.com/api/v3/coins/markets"
api_params = {
    "vs_currency": "usd",
    "order": "market_cap_desc",
    "per_page": 50,
    "page": 1,
    "sparkline": False
}

# Function to fetch cryptocurrency data from the API
def fetch_crypto_data():
    print("Fetching live cryptocurrency data...")
    try:
        api_response = requests.get(api_url, params=api_params)
        api_response.raise_for_status()  # Raise an error for bad responses
        crypto_data = api_response.json()
        data_frame = pd.DataFrame(crypto_data, columns=[
            "name", "symbol", "current_price", "market_cap", "total_volume", "price_change_percentage_24h"
        ])
        data_frame["price_change_percentage_24h"] = data_frame["price_change_percentage_24h"].round(2)  # Round percentage values
        return data_frame
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return pd.DataFrame()

# Function to analyze the fetched data
def analyze_crypto_data(data_frame):
    print("Analyzing cryptocurrency data...")
    top_5_cryptos = data_frame.nlargest(5, "market_cap")
    avg_price = data_frame["current_price"].mean()
    highest_change = data_frame.nlargest(1, "price_change_percentage_24h")
    lowest_change = data_frame.nsmallest(1, "price_change_percentage_24h")

    print("\nTop 5 Cryptocurrencies by Market Cap:\n", top_5_cryptos)
    print(f"\nAverage Price of Top 50 Cryptocurrencies: ${avg_price:.2f}")
    print("\nHighest 24-hour Change:\n", highest_change)
    print("\nLowest 24-hour Change:\n", lowest_change)

# Function to check if the Excel file is open and handle permission issues
def check_file_availability(filename):
    if os.path.exists(filename):
        try:
            os.rename(filename, filename + ".temp")
        except OSError:
            print("The file is currently open. Please close it and try again.")
            return False
    return True

# Function to write the data to an Excel file with formatting and chart
def write_to_excel(data_frame, filename="crypto_data.xlsx"):
    if not check_file_availability(filename):
        return

    try:
        data_frame.to_excel(filename, index=False, sheet_name="Crypto Data")
        workbook = load_workbook(filename)
        worksheet = workbook["Crypto Data"]

        # to format headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # to adjust column widths
        column_widths = {"A": 20, "B": 10, "C": 15, "D": 20, "E": 20, "F": 30}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # for applying conditional formatting for 24-hour percentage change
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6):
            for cell in row:
                if cell.value > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # or adding bar chart for market cap distribution
        chart = BarChart()
        chart.title = "Market Cap Distribution"
        chart.x_axis.title = "Cryptocurrency"
        chart.y_axis.title = "Market Cap (USD)"
        data = Reference(worksheet, min_col=4, min_row=1, max_row=worksheet.max_row, max_col=4)
        labels = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        worksheet.add_chart(chart, "H2")

        workbook.save(filename)
        print("Data written and formatted in Excel successfully!")
    except Exception as e:
        print(f"Failed to write to Excel: {e}")

# Main function to fetch, analyze, and update the Excel file every 5 minutes
def main():
    excel_filename = "/home/deepanshuverma/crypto_data.xlsx"
    while True:
        crypto_data_frame = fetch_crypto_data()
        if not crypto_data_frame.empty:
            analyze_crypto_data(crypto_data_frame)
            write_to_excel(crypto_data_frame, excel_filename)
        else:
            print("No data fetched. Retrying in 5 minutes...")
        
        print("Waiting for 5 minutes before the next update...")
        time.sleep(300)  # Sleep for 5 minutes (300 seconds)

if __name__ == "__main__":
    main()

excel_filename = "C:/Users/Deep/OneDrive/Documents/crypto_data.xlsx" 
